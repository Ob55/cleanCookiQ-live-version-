#!/usr/bin/env python3
"""
One-shot loader for the KoBo "Clean Cooking for Sustainable Financing" export.

  python scripts/import_kobo_institutions.py \
    --file "Clean Cooking for Sustainable Financing_300 schools (3).xlsx"

Idempotent: upserts on `kobo_submission_uuid`. Also archives the full 174-column
raw row in `kobo_submissions_raw` so no survey data is lost.

Uses the same Supabase Management API pattern as `apply_migrations.py` so the
service-role key is never shipped around.
"""

import argparse
import json
import os
import sys
import urllib.request
import urllib.error
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required.  pip install openpyxl")

TOKEN = os.environ.get("SUPABASE_MGMT_TOKEN", "")
if not TOKEN:
    sys.exit("Error: set SUPABASE_MGMT_TOKEN env var before running this script.")
PROJECT_REF = os.environ.get("SUPABASE_PROJECT_REF", "bnbhattryqbterblybzw")
API_URL = f"https://api.supabase.com/v1/projects/{PROJECT_REF}/database/query"

# ---------- column indices (1-based, matches openpyxl) ----------
COL = {
    "lat": 6, "lon": 7,
    "inst_type": 13, "name": 14, "school_class": 15,
    "county": 16,
    "subcounty_a": 17, "subcounty_b": 18, "subcounty_c": 19, "subcounty_d": 20,
    "phone": 23,
    "students": 24, "staff": 28,
    "day_board": 35, "has_kitchen": 36,
    "meals_per_day": 38, "meal_cost": 40,
    "main_fuel": 67, "fuel_sourcing": 68,
    "firewood_kg_day": 55, "charcoal_kg_day": 56, "lpg_kg_day": 63,
    "electricity_kwh_month": 64, "biomass_pellets_kg_day": 60,
    "grid": 98, "outages": 100,
    "kobo_id": 164, "kobo_uuid": 165,
}

# ---------- fuel pricing (KES) — update when rates move ----------
FUEL_PRICE_KES = {
    "firewood":        25,   # per kg
    "charcoal":        55,   # per kg
    "lpg":            260,   # per kg
    "electric":        27,   # per kWh (already monthly)
    "biomass_pellets": 30,   # per kg
}

INST_TYPE_MAP = {
    "school":    "school",
    "church":    "faith_based",
    "hospital":  "hospital",
    "prison":    "prison",
    "factory":   "factory",
    "hotel":     "hotel",
    "restaurant":"restaurant",
}

FUEL_MAP = {
    "firewood":        "firewood",
    "charcoal":        "charcoal",
    "lpg":             "lpg",
    "biogas":          "biogas",
    "electricity":     "electric",
    "electric":        "electric",
    "biomas pellets":  "other",   # enum has no pellets value yet
    "biomass pellets": "other",
    "ethanol":         "other",
    "ethanol/alcohol": "other",
    "ethanol gel":     "other",
    "coal":            "other",
    "kerosene":        "other",
    "dung cake":       "other",
    "crop residue":    "other",
}


def yesno(v):
    if v is None: return None
    return str(v).strip().lower() in ("yes", "true", "1", "y")


def clean_str(v):
    if v is None: return None
    s = str(v).strip()
    return s or None


_NUM_SENTINELS = {"", "n/a", "na", "none", "null", "-", "--"}


def to_int(v):
    if v is None: return None
    if isinstance(v, (int, float)): return int(v)
    s = str(v).strip().lower()
    if s in _NUM_SENTINELS: return None
    try:    return int(float(s))
    except (TypeError, ValueError): return None


def to_float(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().lower()
    if s in _NUM_SENTINELS: return None
    try:    return float(s)
    except (TypeError, ValueError): return None


def norm_phone(v):
    if v is None: return None
    digits = "".join(c for c in str(v) if c.isdigit())
    if not digits: return None
    # Kenya numbers: 9 digits after country code (2547xxxxxxxx or 07xxxxxxxx)
    if digits.startswith("254"):   digits = digits
    elif digits.startswith("0"):   digits = "254" + digits[1:]
    elif len(digits) == 9:         digits = "254" + digits
    return "+" + digits


def compute_monthly_fuel_spend(fuel_enum, row_vals):
    """Return KES/month estimate given normalised fuel + raw row values (by COL key)."""
    price = FUEL_PRICE_KES.get(fuel_enum)
    if price is None: return None

    # Electricity survey value is already kWh/month; others are kg/day.
    if fuel_enum == "electric":
        kwh = row_vals.get("electricity_kwh_month")
        if kwh is None: return None
        return round(float(kwh) * price, 2)

    amount_map = {
        "firewood":        "firewood_kg_day",
        "charcoal":        "charcoal_kg_day",
        "lpg":             "lpg_kg_day",
        "other":           None,
    }
    key = amount_map.get(fuel_enum)
    if key is None: return None
    kg_per_day = row_vals.get(key)
    if kg_per_day is None: return None
    return round(float(kg_per_day) * 30 * price, 2)


def run_sql(sql: str):
    payload = json.dumps({"query": sql}).encode()
    req = urllib.request.Request(
        API_URL,
        data=payload,
        headers={
            "Authorization": f"Bearer {TOKEN}",
            "Content-Type": "application/json",
            "User-Agent": "python-urllib/3.10",
            "Accept": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            return resp.status, resp.read().decode()
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()


def sql_lit(v):
    """Render a Python value as a Postgres literal."""
    if v is None: return "NULL"
    if isinstance(v, bool): return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)): return repr(v)
    s = str(v).replace("'", "''")
    return f"'{s}'"


def build_payload(ws, row):
    """Extract one row -> (mapped_institution_dict, raw_payload_dict)."""
    get = lambda key: ws.cell(row=row, column=COL[key]).value

    raw = {}
    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers, 1):
        if h is None: continue
        v = ws.cell(row=row, column=i).value
        raw[str(h)] = v.isoformat() if hasattr(v, "isoformat") else v

    name = clean_str(get("name"))
    lat = get("lat"); lon = get("lon")
    inst_type_raw = (clean_str(get("inst_type")) or "").lower()
    main_fuel_raw = (clean_str(get("main_fuel")) or "").lower()

    inst_type = INST_TYPE_MAP.get(inst_type_raw, "other")
    current_fuel = FUEL_MAP.get(main_fuel_raw)  # None ok -> leaves column null

    subcounty = (
        clean_str(get("subcounty_a"))
        or clean_str(get("subcounty_b"))
        or clean_str(get("subcounty_c"))
        or clean_str(get("subcounty_d"))
    )

    row_vals = {
        "firewood_kg_day":        to_float(get("firewood_kg_day")),
        "charcoal_kg_day":        to_float(get("charcoal_kg_day")),
        "lpg_kg_day":             to_float(get("lpg_kg_day")),
        "electricity_kwh_month":  to_float(get("electricity_kwh_month")),
        "biomass_pellets_kg_day": to_float(get("biomass_pellets_kg_day")),
    }
    monthly_spend = compute_monthly_fuel_spend(current_fuel, row_vals)

    meals = to_int(get("meals_per_day"))
    inst = {
        "name":                   name,
        "latitude":               to_float(lat),
        "longitude":              to_float(lon),
        "institution_type":       inst_type,
        "county":                 clean_str(get("county")) or "Unknown",
        "sub_county":             subcounty,
        "contact_phone":          norm_phone(get("phone")),
        "number_of_students":     to_int(get("students")),
        "number_of_staff":        to_int(get("staff")),
        "school_type":            clean_str(get("day_board")),
        "has_dedicated_kitchen":  yesno(get("has_kitchen")),
        "meals_per_day":          meals,
        "meals_served_per_day":   meals,
        "avg_meal_cost_ksh":      to_float(get("meal_cost")),
        "current_fuel":           current_fuel,
        "fuel_sourcing":          clean_str(get("fuel_sourcing")),
        "monthly_fuel_spend":     monthly_spend,
        "ownership_type":         clean_str(get("school_class")),
        "grid_connected":         yesno(get("grid")),
        "outages_per_month":      clean_str(get("outages")),
        "pipeline_stage":         "identified",
        "kobo_submission_id":     to_int(get("kobo_id")),
        "kobo_submission_uuid":   clean_str(get("kobo_uuid")),
    }
    return inst, raw


INST_COLS = [
    "name", "latitude", "longitude", "institution_type", "county", "sub_county",
    "contact_phone", "number_of_students", "number_of_staff", "school_type",
    "has_dedicated_kitchen", "meals_per_day", "meals_served_per_day",
    "avg_meal_cost_ksh", "current_fuel", "fuel_sourcing", "monthly_fuel_spend",
    "ownership_type", "grid_connected", "outages_per_month", "pipeline_stage",
    "kobo_submission_id", "kobo_submission_uuid",
]


def build_upsert_sql(rows):
    values = []
    for inst, _ in rows:
        values.append("(" + ", ".join(sql_lit(inst.get(c)) for c in INST_COLS) + ")")
    update_clause = ", ".join(f"{c}=EXCLUDED.{c}" for c in INST_COLS if c != "kobo_submission_uuid")
    return (
        f"INSERT INTO public.institutions ({', '.join(INST_COLS)}) VALUES\n  "
        + ",\n  ".join(values)
        + f"\n  ON CONFLICT (kobo_submission_uuid) WHERE kobo_submission_uuid IS NOT NULL\n"
        + f"  DO UPDATE SET {update_clause}\n"
        + f"  RETURNING id, kobo_submission_uuid;"
    )


def build_raw_upsert_sql(rows):
    values = []
    for inst, raw in rows:
        uuid = inst.get("kobo_submission_uuid")
        sid = inst.get("kobo_submission_id")
        payload_json = json.dumps(raw, default=str).replace("'", "''")
        values.append(
            f"({sql_lit(sid)}, {sql_lit(uuid)}, '{payload_json}'::jsonb)"
        )
    return (
        "INSERT INTO public.kobo_submissions_raw (kobo_submission_id, kobo_uuid, payload) VALUES\n  "
        + ",\n  ".join(values)
        + "\n  ON CONFLICT (kobo_uuid) DO UPDATE\n"
        + "  SET payload = EXCLUDED.payload, imported_at = now();"
    )


def chunked(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i : i + n]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="Path to KoBo .xlsx export")
    ap.add_argument("--dry-run", action="store_true", help="Parse + validate only — do not touch Supabase")
    ap.add_argument("--chunk", type=int, default=50, help="Rows per upsert statement")
    args = ap.parse_args()

    path = Path(args.file)
    if not path.exists():
        sys.exit(f"File not found: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    parsed = []
    skipped = []
    for r in range(2, ws.max_row + 1):
        # Skip blank rows (no name *and* no coords)
        name = ws.cell(row=r, column=COL["name"]).value
        lat = ws.cell(row=r, column=COL["lat"]).value
        lon = ws.cell(row=r, column=COL["lon"]).value
        if not name and lat is None and lon is None:
            continue
        try:
            inst, raw = build_payload(ws, r)
            if not inst["name"]:
                skipped.append((r, "missing name")); continue
            if inst["latitude"] is None or inst["longitude"] is None:
                skipped.append((r, "missing coords")); continue
            parsed.append((inst, raw))
        except Exception as e:
            skipped.append((r, f"parse error: {e}"))

    print(f"Parsed  : {len(parsed)} rows")
    print(f"Skipped : {len(skipped)} rows")
    for r, why in skipped[:10]:
        print(f"  row {r}: {why}")
    if len(skipped) > 10:
        print(f"  …and {len(skipped) - 10} more")

    # Mapping sanity report
    type_counts = {}
    fuel_counts = {}
    spend_have = 0
    for inst, _ in parsed:
        type_counts[inst["institution_type"]] = type_counts.get(inst["institution_type"], 0) + 1
        fuel_counts[inst["current_fuel"] or "(none)"] = fuel_counts.get(inst["current_fuel"] or "(none)", 0) + 1
        if inst["monthly_fuel_spend"] is not None:
            spend_have += 1
    print("\ninstitution_type ->", type_counts)
    print("current_fuel   ->", fuel_counts)
    print(f"monthly_fuel_spend populated on {spend_have}/{len(parsed)} rows")

    if args.dry_run:
        print("\n--dry-run: stopping before DB calls.")
        return

    if not parsed:
        sys.exit("Nothing to import.")

    # Push institutions
    for batch in chunked(parsed, args.chunk):
        status, body = run_sql(build_upsert_sql(batch))
        if status >= 300:
            print(f"FAIL institutions upsert  HTTP {status}")
            print(body[:2000]); sys.exit(1)
    print(f"\n✓ {len(parsed)} rows upserted into institutions")

    # Push raw payloads
    for batch in chunked(parsed, args.chunk):
        status, body = run_sql(build_raw_upsert_sql(batch))
        if status >= 300:
            print(f"FAIL kobo_submissions_raw upsert  HTTP {status}")
            print(body[:2000]); sys.exit(1)
    print(f"✓ {len(parsed)} rows upserted into kobo_submissions_raw")

    # Link raw rows back to institutions.id
    link_sql = """
      UPDATE public.kobo_submissions_raw r
      SET institution_id = i.id
      FROM public.institutions i
      WHERE r.kobo_uuid = i.kobo_submission_uuid
        AND r.institution_id IS DISTINCT FROM i.id;
    """
    status, body = run_sql(link_sql)
    if status >= 300:
        print(f"FAIL raw->institution link  HTTP {status}")
        print(body[:2000]); sys.exit(1)
    print("✓ raw payloads linked to institution rows")
    print("\nDone.")


if __name__ == "__main__":
    main()
